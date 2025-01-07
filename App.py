import openpyxl as xl

wb = xl.load_workbook("Employee_Workbook.xlsx")
sheet = wb["Sheet1"]

def find_employee(sheet, name):
      
    for row in range(2, sheet.max_row + 1):
       cell = sheet.cell(row, 1)
       if(name == cell.value):
            return row

    return -1


def add_employee(name, role):

    cell = sheet.cell(sheet.max_row + 1, 1)
    cell.value = name
    cell = sheet.cell(sheet.max_row, 2)
    cell.value = role
    print("\nWelcome to the team :)\n")


def view_shift(sheet, employee_row):
    print("------ Schedule for this week ------")
    
    for i in range(3, sheet.max_column + 1):
        day = sheet.cell(1, i)
        time = sheet.cell(employee_row, i)
        if not time.value == None:
            print(f"{day.value:<10}: {time.value}")


def manager_tools(sheet, Employeerow):
     while True:
            choice = input("""\nWhat would you like to do:
1. View my shifts
2. Add Shifts
3. Quit
""")
            
            if choice == "1":
                view_shift(sheet, Employeerow)
            elif choice == "2":
                pass
            elif choice == "3":
                break
            else:
                print("Enter a correct input!\n")


def associate_tools(sheet, Employeerow):
    while True:
        choice = input("""\nWhat would you like to do:
1. View my shifts
2. Quit
""")
        
        if choice == "1":
            view_shift(sheet, Employeerow)
        elif choice == "2":
            break
        else:
            print("Enter a correct input!\n")

# Main Program

print("##### WELCOME TO THE SHIFT APP #####")
name = input("Please enter your name: ")

if find_employee(sheet, name) == -1:
    print(f"{name} looks like you are new here")
    role = input("Tell us your role: ")
    add_employee(name, role)

Employeerow = find_employee(sheet, name)
role = sheet.cell(Employeerow, 2).value

if role == "Manager":
    # Manager tools
    manager_tools(sheet, Employeerow)

else:
    associate_tools(sheet, Employeerow)
    

wb.save("Employee_Workbook.xlsx")